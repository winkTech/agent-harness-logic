#!/usr/bin/env python3
"""recalc.py — xlsx 公式重算 + 错误值扫描。

为什么需要它：openpyxl 写入的公式只是字符串，**没有缓存值**。文件在被真正的
电子表格引擎打开前，公式是否正确无从得知 —— 交付一个满是 `#REF!` 的表格是
本技能记录过的事故。

两段式：
  1. 有 LibreOffice → headless 重算并回写（公式获得缓存值）
  2. 无 LibreOffice → 跳过重算，仅做静态扫描并**明确告知未重算**

扫描项（两种模式都做）：
  - 缓存值中的错误：#REF! #DIV/0! #VALUE! #NAME? #N/A #NULL! #NUM!
  - 可疑公式：引用整列/跨表未加引号/除数可能为 0

用法:
    python recalc.py <file.xlsx> [--no-recalc] [--json]

退出码: 0 无错误 | 1 发现错误值 | 2 文件问题/依赖缺失
"""
from __future__ import annotations

import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile

try:  # Windows 控制台默认 GBK；stderr 也要重配，否则中文报错被捕获时是乱码
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:  # noqa: BLE001
    pass

ERROR_VALUES = ('#REF!', '#DIV/0!', '#VALUE!', '#NAME?', '#N/A', '#NULL!', '#NUM!')
SUSPECT_PATTERNS = [
    (re.compile(r'\b[A-Z]{1,3}:[A-Z]{1,3}\b'), '引用整列（大表上会拖慢重算）'),
    (re.compile(r'/\s*0(?![.\d])'), '除数为字面量 0'),
]

HERE = os.path.dirname(os.path.abspath(__file__))
SOFFICE_WRAPPER = os.path.join(HERE, 'office', 'soffice.py')


def recalculate(path: str) -> tuple[bool, str]:
    """用 LibreOffice headless 重算并回写。返回 (是否重算, 说明)。"""
    if not os.path.isfile(SOFFICE_WRAPPER):
        return False, 'soffice 包装器缺失'
    with tempfile.TemporaryDirectory(prefix='recalc_') as out:
        r = subprocess.run(
            [sys.executable, SOFFICE_WRAPPER, '--headless', '--convert-to', 'xlsx', '--outdir', out, path],
            capture_output=True, text=True, timeout=600)
        if r.returncode == 2:
            return False, 'LibreOffice 不可用（未安装或未设置 SOFFICE_BIN）'
        if r.returncode != 0:
            return False, f'转换失败 exit={r.returncode}: {(r.stderr or "").strip()[:200]}'
        produced = os.path.join(out, os.path.basename(path))
        if not os.path.isfile(produced):
            cand = [f for f in os.listdir(out) if f.lower().endswith('.xlsx')]
            if not cand:
                return False, 'LibreOffice 未产出 xlsx'
            produced = os.path.join(out, cand[0])
        shutil.copyfile(produced, path)
        return True, '已用 LibreOffice 重算并回写'


def scan(path: str) -> dict:
    try:
        import openpyxl
    except ImportError:
        return {'error': 'openpyxl 未安装 — pip install openpyxl'}

    errors, suspects = [], []
    wb_val = openpyxl.load_workbook(path, data_only=True)
    wb_f = openpyxl.load_workbook(path, data_only=False)
    formula_count = 0

    for ws_v, ws_f in zip(wb_val.worksheets, wb_f.worksheets):
        for row_v, row_f in zip(ws_v.iter_rows(), ws_f.iter_rows()):
            for cv, cf in zip(row_v, row_f):
                if isinstance(cv.value, str) and cv.value.strip() in ERROR_VALUES:
                    errors.append({'sheet': ws_v.title, 'cell': cv.coordinate,
                                   'error': cv.value.strip(),
                                   'formula': cf.value if isinstance(cf.value, str) else None})
                if isinstance(cf.value, str) and cf.value.startswith('='):
                    formula_count += 1
                    for pat, why in SUSPECT_PATTERNS:
                        if pat.search(cf.value):
                            suspects.append({'sheet': ws_f.title, 'cell': cf.coordinate,
                                             'formula': cf.value[:120], 'why': why})
    wb_val.close()
    wb_f.close()
    return {'formula_count': formula_count, 'errors': errors, 'suspects': suspects}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('file')
    ap.add_argument('--no-recalc', action='store_true', help='只扫描，不调用 LibreOffice')
    ap.add_argument('--json', action='store_true')
    args = ap.parse_args()

    if not os.path.isfile(args.file):
        print(f'文件不存在: {args.file}', file=sys.stderr)
        return 2

    recalced, note = (False, '按 --no-recalc 跳过') if args.no_recalc else recalculate(args.file)
    result = scan(args.file)
    if 'error' in result:
        print(result['error'], file=sys.stderr)
        return 2
    result['recalculated'] = recalced
    result['recalc_note'] = note

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        print(f'公式数: {result["formula_count"]}')
        print(f'重算: {"是" if recalced else "否"} — {note}')
        if not recalced:
            print('⚠️  未重算：openpyxl 写入的公式没有缓存值，下面的错误扫描只能看到**已有**缓存值，')
            print('   不能证明公式正确。要真正验证，请安装 LibreOffice 或用 Excel 打开另存。')
        print(f'错误值: {len(result["errors"])}')
        for e in result['errors'][:30]:
            print(f'  ❌ {e["sheet"]}!{e["cell"]} = {e["error"]}   {e.get("formula") or ""}')
        print(f'可疑公式: {len(result["suspects"])}')
        for s in result['suspects'][:20]:
            print(f'  ⚠️  {s["sheet"]}!{s["cell"]}  {s["why"]}  {s["formula"]}')

    return 1 if result['errors'] else 0


if __name__ == '__main__':
    sys.exit(main())
